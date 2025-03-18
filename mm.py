import asyncio
import emoji
from openpyxl import load_workbook
import json
import logging
import pandas as pd
from openpyxl.workbook import Workbook
from telegram import (
    InlineKeyboardButton, InlineKeyboardMarkup, Update, ReplyKeyboardMarkup, KeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler,
    CallbackContext
)
from datetime import datetime, time, timedelta, date
from yookassa import Configuration, Payment, payment
import uuid
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configuration
TOKEN = os.getenv('TELEGRAM_TOKEN')
YOOKASSA_ACCOUNT_ID = os.getenv('YOOKASSA_ACCOUNT_ID')
YOOKASSA_SECRET_KEY = os.getenv('YOOKASSA_SECRET_KEY')
CARD_NUMBER = os.getenv('CARD_NUMBER')

# YooKassa settings
Configuration.configure(account_id=YOOKASSA_ACCOUNT_ID, secret_key=YOOKASSA_SECRET_KEY)

# Constants
DATA_FILE = "Data.json"
ORDERS = "Заказы.xlsx"
MENU = "https://docs.google.com/spreadsheets/d/1eEEHGwtSV2znQDGJcgGVEQ2PzNTLoDPOT-9vtyQCoQY/export?format=csv"
ADDRESSES_FILE = "Addresses.json"
ORDERS_JSON = "Orders.json"

#logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

CHOOSE_ADDRESS, ENTER_NAME, BROADCAST_MESSAGE, ADD_ADDRESS, ENTER_PHONE, SELECT_ROLE, ENTER_COMMENT = range(7)

def load_data(file_path, default):
    try:
        if not os.path.exists(file_path):
            logger.warning(f"File {file_path} does not exist, using default value")
            return default
            
        with open(file_path, "r", encoding="utf-8") as file:
            data = json.load(file)
            return data
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error in {file_path}: {e}")
        return default
    except Exception as e:
        logger.error(f"Error loading data from {file_path}: {e}")
        return default

def save_data(file_path, data):
    try:
        with open(file_path, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Error saving data to {file_path}: {e}")
        raise

def load_user_data():
    return load_data(DATA_FILE, {"users": []})

def save_user_data(data):
    save_data(DATA_FILE, data)

def load_addresses():
    return load_data(ADDRESSES_FILE, {"addresses": []})

def save_addresses(data):
    save_data(ADDRESSES_FILE, data)

def load_menu_data():
    try:
        if not os.path.exists(MENU):
            logger.error("Menu file does not exist")
            return None
            
        df = pd.read_csv(MENU)
        return df
    except Exception as e:
        logger.error(f"Error loading menu: {e}")
        return None

def normalize_phone_number(phone_number):
    try:
        if not phone_number:
            return None
            
        digits = ''.join(filter(str.isdigit, phone_number))
        if not digits:
            return None
            
        if len(digits) == 11 and digits.startswith('8'):
            return '7' + digits[1:]
        elif len(digits) == 10 and digits.startswith('9'):
            return '7' + digits
        elif len(digits) == 11 and digits.startswith('7'):
            return digits
        elif len(digits) == 12 and digits.startswith('+7'):
            return digits
            
        logger.warning(f"Invalid phone number format: {phone_number}")
        return None
    except Exception as e:
        logger.error(f"Error normalizing phone number: {e}")
        return None

async def under_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = context.user_data.get("phone verified")
    if phone is None:
        await update.message.reply_text(
            "Пожалуйста, подтвердите согласие на обработку ПД \n https://telegra.ph/Soglasie-obrabotki-PD-02-10",
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton("Я согласен ✔")]],
                resize_keyboard=True, one_time_keyboard=True
            )
        )
    else:
        await start(update, context)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        load_user_data()
        user_data = load_user_data()
        chat_id = update.message.chat_id
        if context.user_data.get("phone_verified"):
            user = next((u for u in user_data["users"] if u.get("chat_id") == chat_id), None)
            if user:
                keyboard = get_role_keyboard(user.get("role", "Заказчик"))
                await update.message.reply_text(
                    f"Добро пожаловать, {user['name']}! Ваша роль 🙋‍♂️: {user['role']}.",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
                return

        user = next((u for u in user_data["users"] if u.get("chat_id") == chat_id), None)
        if user:
            context.user_data["phone_verified"] = True
            context.user_data["phone_number"] = user["phone"]
            context.user_data["role"] = user.get("role", "Заказчик")

            logger.info(f"Пользователь уже зарегистрирован: {user['name']}, роль: {user['role']}")

            keyboard = get_role_keyboard(user.get("role", "Заказчик"))
            await update.message.reply_text(
                f"Добро пожаловать, {user['name']}! Ваша роль: {user['role']}.",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return

        contact = update.message.contact
        if contact:
            phone_number = normalize_phone_number(contact.phone_number)
            user = next((u for u in user_data["users"] if u["phone"] == phone_number), None)

            if user:
                context.user_data["phone_verified"] = True
                context.user_data["phone_number"] = phone_number
                context.user_data["role"] = user.get("role", "Заказчик")

                logger.info(f"Роль пользователя: {context.user_data.get('role')}")

                keyboard = get_role_keyboard(user.get("role", "Заказчик"))
                await update.message.reply_text(
                    f"Добро пожаловать, {user['name']}! Ваша роль 🙋‍♂️: {user['role']}.",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
            else:
                addresses = load_addresses().get("addresses", [])
                if not addresses:
                    await update.message.reply_text(
                        "Список адресов доставки недоступен. Свяжитесь с администратором."
                    )
                    return

                context.user_data["phone_number"] = phone_number
                keyboard = []
                for address in addresses:
                    # Convert address to string and create a unique callback data
                    address_str = str(address)[:64]  # Limit to 64 characters
                    callback_data = f"addr_{hash(address_str) % 10000}"  # Create shorter unique identifier
                    keyboard.append([InlineKeyboardButton(address_str, callback_data=callback_data)])
                
                # Store address mapping in context
                context.user_data["address_mapping"] = {
                    f"addr_{hash(str(addr))% 10000}": str(addr) 
                    for addr in addresses
                }
                
                await update.message.reply_text(
                    "Выберите адрес доставки 🏘:",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
                return CHOOSE_ADDRESS
        else:
            await update.message.reply_text(
                "Пожалуйста, подтвердите ваш номер телефона.",
                reply_markup=ReplyKeyboardMarkup(
                    [[KeyboardButton("Подтвердить номер телефона", request_contact=True)]],
                    resize_keyboard=True, one_time_keyboard=True
                )
            )
    except Exception as e:
        logger.error(f"Ошибка в команде start: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    payment_id = query.data.split("_")[2]

    try:
        payment = Payment.find_one(payment_id)
        status = payment.status
        await query.edit_message_text(f'Статус платежа {payment_id}: {status}')
    except Exception as e:
        await query.edit_message_text(f'Ошибка при проверке статуса платежа: {str(e)}')

def get_role_keyboard(role):
    if role == "Администратор":
        return [["Список заказов", "Сообщить всем"], ["Добавить адрес доставки", "Выгрузка заказов"]]
    elif role == "Заказчик":
        return [["Сделать заказ 🍴", "Корзина 🗑"]]

async def choose_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        query = update.callback_query
        await query.answer()

        # Get the actual address from the mapping
        callback_data = query.data
        address = context.user_data["address_mapping"].get(callback_data)
        
        if not address:
            await query.edit_message_text("Ошибка: адрес не найден")
            return

        phone_number = context.user_data.get("phone_number")
        if not phone_number:
            await query.edit_message_text("Ошибка регистрации. Попробуйте ещё раз.")
            return

        context.user_data["address"] = address
        await query.edit_message_text(f"Адрес доставки выбран: {address}. Введите ваше имя и фамилию:")
        return ENTER_NAME
    except Exception as e:
        logger.error(f"Error in choose_address: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def enter_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        name = update.message.text.strip()
        if not name:
            await update.message.reply_text("Имя не может быть пустым. Пожалуйста, введите ваше имя и фамилию:")
            return ENTER_NAME

        phone_number = context.user_data.get("phone_number")
        address = context.user_data.get("address")

        if not phone_number or not address:
            logger.error("Missing phone number or address in user data")
            await update.message.reply_text("Ошибка регистрации. Попробуйте снова.")
            return ConversationHandler.END

        user_data = load_user_data()
        user_data["users"].append({
            "phone": phone_number,
            "role": "Заказчик",
            "address": address,
            "name": name,
            "chat_id": update.message.chat_id
        })
        save_user_data(user_data)

        await update.message.reply_text(f"Регистрация завершена. Добро пожаловать, {name}!")
        role = context.user_data.get("role", "Заказчик")
        
        keyboard = get_role_keyboard(role)
        await update.message.reply_text(
            f"Теперь вы можете заказывать, {name}!",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"Error in enter_name: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")
        return ConversationHandler.END
    
async def show_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    today = datetime.now()
    days = [today + timedelta(days=i) for i in range(7)]
    cutoff_time = time(20, 00) #ТУТ МЕНЯТЬ ВРЕМЯ 10 - ЧАСЫ; 00 - МИНУТЫ!!!!!!!!!!!!!!!!!!!!!!!!!

    keyboard = []
    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    for day in days:
        if day.date() == today.date() and datetime.now().time() >= cutoff_time:
            continue
        day_name = days_of_week[day.weekday()]
        button_text = f"{day.strftime('%d.%m.%Y')} ({day_name})"
        keyboard.append([InlineKeyboardButton(button_text, callback_data=day.strftime('%d.%m.%Y'))])


    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите дату 📆:", reply_markup=reply_markup)

async def handle_menu_and_lunch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if isinstance(update, Update) and update.callback_query:
        query = update.callback_query
        selected_date_str = query.data
        selected_date_full = datetime.strptime(selected_date_str, '%d.%m.%Y')
        days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
        day_index = selected_date_full.weekday()
        selected_day_name = days_of_week[day_index]

        await query.answer()
        await query.edit_message_text(f"Вы выбрали дату 📆: {selected_date_str} ({selected_day_name})")
        context.user_data["selected_date"] = selected_date_str
        context.user_data["selected_day_name"] = selected_day_name

        try:
            menu_data = pd.read_csv(MENU)
            menu_data['Цена'] = menu_data['Цена'].astype(str) + ' рублей'

            week_number = selected_date_full.isocalendar()[1] % 2

            daily_menu = menu_data[(menu_data['День недели'] == selected_day_name) & (menu_data['Неделя'] == week_number)]

            if daily_menu.empty:
                await query.message.reply_text("К сожалению, на эту дату нет меню.")
                return

            lunch_items = daily_menu.groupby('Название').agg({'Блюдо': list, 'Цена': 'first'}).reset_index()

            menu_text = f"Меню на {selected_date_str} ({days_of_week[day_index]})\n\n"

            for index, row in lunch_items.iterrows():
                menu_text += f"*{row['Название']}* ({row['Цена']}):\n"
                for i, dish in enumerate(row['Блюдо']):
                    menu_text += f"{i+1}. {dish}\n"
                menu_text += "\n"

            await query.message.reply_text(menu_text)

            complex_lunches = daily_menu[daily_menu['Название'] == 'Комплексный обед']['Название'].unique().tolist()
            drinks = daily_menu[daily_menu['Название'] == 'Напиток']['Блюдо'].unique().tolist()
            salads = daily_menu[daily_menu['Название'] == 'Салат']['Блюдо'].unique().tolist()

            keyboard = []
            if complex_lunches:
                row = [KeyboardButton(option) for option in complex_lunches]
                keyboard.append(row)

            if drinks:
                row = [KeyboardButton(option) for option in drinks]
                keyboard.append(row)

            if salads:
                row = [KeyboardButton(option) for option in salads]
                keyboard.append(row)

            keyboard.append([KeyboardButton("Назад 🔙")])
            keyboard.append([KeyboardButton("Корзина 🗑")])
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
            await query.message.reply_text("Выберите обед 🍜:", reply_markup=reply_markup)
        except Exception as e:
            await query.message.reply_text(f"Ошибка при загрузке меню: {e}")
            return

    elif isinstance(update, Update) and update.message and update.message.text:
        message = update.message.text
        phone = context.user_data.get("phone_number")
        if phone is None:
            await update.message.reply_text("Ваш номер телефона не зарегистрирован, перезапустите бота!")
            return
        
        selected_date = context.user_data.get("selected_date")
        selected_day_name = context.user_data.get("selected_day_name")

        if selected_date is None:
            await update.message.reply_text("Выберите дату, прежде чем заказывать обед.")
            return

        try:
            menu_data = pd.read_csv(MENU)
            daily_menu = menu_data[menu_data['День недели'] == selected_day_name] 

            if message in daily_menu['Название'].unique():
                complex_lunch_options = daily_menu[daily_menu['Название'] == message]
                if not complex_lunch_options.empty:
                    price = complex_lunch_options['Цена'].iloc[0]
                else:
                    await update.message.reply_text(f"Цена для {message} не найдена в меню.")
                    return

            else:
                price_row = daily_menu[daily_menu['Блюдо'] == message]
                if not price_row.empty:
                    price = price_row['Цена'].iloc[0]
                else:
                    await update.message.reply_text(f"Цена для {message} не найдена в меню.")
                    return


            try:
                with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                    try:
                        orders = json.load(f)

                        if not isinstance(orders, list):
                            orders = []
                    except json.JSONDecodeError:

                        orders = []
            except FileNotFoundError:

                orders = []


            new_order = {
                "Номер телефона": phone,
                "Дата": selected_date,
                "День недели": selected_day_name,
                "Обед": message,
                "Цена": int(price),
            }
            try:
                with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                    try:
                        orders = json.load(f)
                        if not isinstance(orders, list):
                            orders = []
                    except json.JSONDecodeError:
                        orders = []
            except FileNotFoundError:
                orders = []

            orders.append(new_order)

            with open(ORDERS_JSON, 'w', encoding='utf-8') as f:
                json.dump(orders, f, ensure_ascii=False, indent=4)


            await update.message.reply_text(f"Ваш выбор ({message}) записан! Цена: {price} рублей.")

            daily_menu = menu_data[menu_data['День недели'] == selected_day_name]
            complex_lunches = daily_menu[daily_menu['Название'] == 'Комплексный обед']['Название'].unique().tolist()
            drinks = daily_menu[daily_menu['Название'] == 'Напиток']['Блюдо'].unique().tolist()
            salads = daily_menu[daily_menu['Название'] == 'Салат']['Блюдо'].unique().tolist()

            keyboard = []
            if complex_lunches:
                row = [KeyboardButton(option) for option in complex_lunches]
                keyboard.append(row)

            if drinks:
                row = [KeyboardButton(option) for option in drinks]
                keyboard.append(row)

            if salads:
                row = [KeyboardButton(option) for option in salads]
                keyboard.append(row)

            keyboard.append([KeyboardButton("Нет, спасибо")])
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
            await update.message.reply_text("Выберите ещё что-нибудь или нажмите 'Нет, спасибо':", reply_markup=reply_markup)

        except Exception as e:
            await update.message.reply_text(f"Ошибка при записи заказа: {e}")
            return

async def move_orders_to_excel(phone, payment_status="Не оплачено", orders_json_path=ORDERS_JSON, orders_excel_path=ORDERS):
    try:
        if not os.path.exists(orders_json_path):
            logger.error(f"Orders JSON file does not exist: {orders_json_path}")
            return False, None

        with open(orders_json_path, "r", encoding="utf-8") as f:
            orders = json.load(f)

        user_orders = [order for order in orders if str(order.get("Номер телефона")).strip() == str(phone).strip()]
        if not user_orders:
            logger.warning(f"No orders found for phone: {phone}")
            return False, []

        order_id = str(uuid.uuid4())

        for order in user_orders:
            order["order_id"] = order_id
            order["Статус оплаты"] = payment_status
            order["Комментарий"] = order.get("Комментарий", "Без комментария")

        try:
            if os.path.exists(orders_excel_path):
                wb = load_workbook(orders_excel_path)
            else:
                wb = Workbook()
                sheet = wb.active
                sheet.append([
                    "Номер телефона", "Дата", "Обед", "Цена", "Статус оплаты",
                    "День недели", "Адрес доставки", "Имя заказчика", "order_id", "Комментарий"
                ])
        except Exception as e:
            logger.error(f"Error handling Excel file: {e}")
            return False, None

        sheet = wb.active
        for order in user_orders:
            sheet.append([
                order.get("Номер телефона", ""),
                order.get("Дата", ""),
                order.get("Обед", ""),
                order.get("Цена", ""),
                order.get("Статус оплаты", ""),
                order.get("День недели", ""),
                order.get("Адрес доставки", ""),
                order.get("Имя заказчика", ""),
                order.get("order_id", ""),
                order.get("Комментарий", "Без комментария")
            ])

        wb.save(orders_excel_path)

        remaining_orders = [order for order in orders if str(order.get("Номер телефона")).strip() != str(phone).strip()]
        with open(orders_json_path, "w", encoding="utf-8") as f:
            json.dump(remaining_orders, f, ensure_ascii=False, indent=4)

        return True, order_id

    except Exception as e:
        logger.error(f"Error moving orders to Excel: {e}")
        return False, None


async def menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if context.user_data.get("role") == "Администратор":
            await update.message.reply_text("У вас нет доступа к этой функции.")
            return

        keyboard = [["Меню", "Мои заказы", "Корзина"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text("Главное меню:", reply_markup=reply_markup)
    except Exception as e:
        logger.error(f"Error in menu: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    role = context.user_data.get("role")
    logger.info(f"Роль пользователя в broadcast_start: {role}")

    if role != "Администратор":
        await update.message.reply_text("У вас нет прав для использования этой функции.")
        return

    await update.message.reply_text("Введите сообщение, которое вы хотите отправить всем пользователям.")
    return BROADCAST_MESSAGE

async def broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        message = update.message.text
        user_data = load_user_data()

        for user in user_data["users"]:
            chat_id = user.get("chat_id")
            if chat_id:
                try:
                    await context.bot.send_message(chat_id=chat_id, text=f"[Сообщение от администратора ✉]\n{message}")
                except Exception as e:
                    logger.error(f"Error sending message to {chat_id}: {e}")

        await update.message.reply_text("Сообщение было отправлено всем пользователям.")
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"Error in broadcast_message: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def add_address_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    role = context.user_data.get("role")
    logger.info(f"Роль пользователя в add_address_start: {role}")

    if role != "Администратор":
        await update.message.reply_text("У вас нет прав для использования этой функции ❌.")
        return

    await update.message.reply_text("Введите адрес, который вы хотите добавить в список доступных для доставки 🏚.")
    return ADD_ADDRESS

async def add_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        address = update.message.text
        addresses = load_addresses()
        addresses["addresses"].append(address)
        save_addresses(addresses)

        await update.message.reply_text(f"Адрес '{address}' был успешно добавлен.")
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"Error in add_address: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def handle_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        text = update.message.text
        logger.info(f"Нажата кнопка: {text}")  # Логируем нажатую кнопку

        if text == "Сделать заказ 🍴":
            await show_menu(update, context)
        elif context.user_data.get("awaiting_comment"):
            await handle_comment(update, context)
            return
        elif text == "Корзина 🗑":
            await show_cart(update, context)
        elif text == "Список заказов":
            await show_all_orders(update, context)
        elif text == "Сообщить всем":
            await broadcast_start(update, context)
        elif text == "Добавить адрес доставки ":
            await add_address_start(update, context)
        elif text == "Комплексный обед":
            await handle_complex_lunch(update, context, "Комплексный обед")
        elif text == "Морс":
            await handle_drink(update, context, "Морс")
        elif text == "Компот":
            await handle_drink(update, context, "Компот")
        elif text == "Цезарь с сёмгой":
            await handle_salad(update, context, "Цезарь с сёмгой")
        elif text == "Цезарь с курицей":
            await handle_salad(update, context, "Цезарь с курицей")
        elif text == "Оплатить картой💳":
            await pay(update, context)
        elif text == "Назад 🔙":
            await show_menu(update, context)
        elif text == "Нет, спасибо":
            await update.message.reply_text("Спасибо за ваш заказ! Если хотите что-то ещё, выберите из меню.")
        elif text.startswith("Заказать на "):
            next_day_str = text.replace("Заказать на ", "")
            context.user_data["selected_date"] = next_day_str
            await show_menu(update, context)
        elif text == "Вернуться в главное меню":
            await show_main_menu(update, context)
        elif text == "Очистить корзину❌":
            await clear_cart(update, context)
        elif text == "Выгрузка заказов":
            await import_excel(update, context)
        elif text == "Оплатить наличными":
            if update.callback_query:
                pass
            else:
                await handle_payment_selection(update, context)
        elif text == "Я согласен ✔":
            await update.message.reply_text("Спасибо за согласие! Переходим к следующему шагу.")
            await start(update, context)

        else:
            await update.message.reply_text("Неизвестная команда. Пожалуйста, выберите действие из меню.")

    except Exception as e:
        logger.error(f"Ошибка при обработке кнопки: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def import_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    role = context.user_data.get("role")
    if role != "Администратор":
        await update.message.reply_text("У вас нет доступа к этой команде")
        return
    await update.message.reply_document(ORDERS)


async def clear_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        phone_number = context.user_data.get("phone_number")
        if not phone_number:
            await update.message.reply_text("Ваш номер телефона не зарегестрирован")
            return
        try:
            with open(ORDERS_JSON, "r", encoding="utf-8") as f:
                orders = json.load(f)
        except FileNotFoundError:
            await update.message.reply_text("Заказов нету")
            return
        except json.JSONDecodeError:
            await update.message.reply_text("Ошибка доступа при обращении к файлу")
            return
        
        initial_count = len(orders)
        orders = [order for order in orders if order.get("Номер телефона") != phone_number]

        with open(ORDERS_JSON, "w", encoding="utf-8") as f:
            json.dump(orders, f, ensure_ascii=False, indent=4)

        if len(orders) < initial_count:
            await update.message.reply_text("Корзина успешно очищена")
            await show_main_menu(update, context)
        else:
            await update.message.reply_text("Корзина пуста")
        
    except Exception as e:
        logger.error(f"Ошибка при очистке корзины: {e}")
        await update.message.reply_text("Ошибка при очистке корзины")

async def handle_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        selected_date = context.user_data.get("selected_date")
        phone_number = context.user_data.get("phone_number")

        if not selected_date or not phone_number:
            await update.message.reply_text("Ошибка: не удалось найти данные о заказе.")
            return

        try:
            orders_df = pd.read_excel(ORDERS)
        except FileNotFoundError:
            await update.message.reply_text("Файл с заказами не найден.")
            return
        phone_number_clean = ''.join(filter(str.isdigit, phone_number))
        orders_df['Номер телефона'] = orders_df['Номер телефона'].astype(str).str.replace('[^0-9]', '', regex=True)

        user_orders = orders_df[
            (orders_df['Номер телефона'] == phone_number_clean) &
            (orders_df['Дата'] == selected_date)
        ]

        if user_orders.empty:
            await update.message.reply_text("Нет заказов для отмены.")
            return
        orders_df = orders_df[
            ~((orders_df['Номер телефона'] == phone_number_clean) &
              (orders_df['Дата'] == selected_date))
        ]
        orders_df.to_excel(ORDERS, index=False)

        await update.message.reply_text("Ваши заказы успешно отменены!")
        await show_main_menu(update, context)

    except Exception as e:
        logger.error(f"Ошибка при отмене заказов: {e}")
        await update.message.reply_text("Произошла ошибка при отмене заказов. Пожалуйста, попробуйте снова.")

async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        role = context.user_data.get("role", "Заказчик")
        if role == "Администратор":
            keyboard = [
                ["Список заказов", "Сообщить всем"],
                ["Добавить адрес доставки", "Импорт chat_id"]
            ]
        else:
            keyboard = [
                ["Сделать заказ 🍴", "Корзина 🗑"]
            ]

        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text("Главное меню:", reply_markup=reply_markup)

    except Exception as e:
        logger.error(f"Ошибка при отображении главного меню: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        text = update.message.text
        await handle_buttons(update, context)
    except Exception as e:
        logger.error(f"Ошибка при обработке текстового сообщения: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def handle_drink(update: Update, context: ContextTypes.DEFAULT_TYPE, drink_name: str):
    user_data = load_user_data()
    pay = context.user_data.get("payment_id")
    try:
            phone = context.user_data.get("phone_number")
            user = next((u for u in user_data["users"] if u["phone"] == phone), None)
            if phone is None:
                await update.message.reply_text("Ваш номер телефона не зарегистрирован, перезапустите бота!")
                return
            selected_date = context.user_data.get("selected_date")
            if selected_date is None:
                await update.message.reply_text("Выберите дату, прежде чем заказывать обед.")
                return
            selected_day_name = context.user_data.get("selected_day_name")
            address = user['address']
            if address is None:
                await update.message.reply_text("Вы не выбрали адрес, перезапустите бота!")
                return

            try:
                menu_data = pd.read_csv(MENU)
                drink_price = dict(zip(menu_data['Блюдо'], menu_data['Цена']))

                price = drink_price.get(drink_name)
                if price is None:
                    await update.message.reply_text(f"Цена для {drink_name} не найдена в меню.")
                    return
                try:
                    with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                        try:
                            orders = json.load(f)

                            if not isinstance(orders, list):
                                orders = []
                        except json.JSONDecodeError:

                            orders = []
                except FileNotFoundError:

                    orders = []

                new_order = {
                "Номер телефона": phone,
                "Дата": selected_date,
                "День недели": selected_day_name,
                "Обед": drink_name,
                "Цена": int(price),
                "Статус оплаты": "Не оплачено",
                "Адрес доставки": address,
                "Имя заказчика": user["name"],
                }
                try:
                    with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                        try:
                            orders = json.load(f)
                            if not isinstance(orders, list):
                                orders = []
                        except json.JSONDecodeError:
                            orders = []
                except FileNotFoundError:
                    orders = []

                orders.append(new_order)

                with open(ORDERS_JSON, 'w', encoding='utf-8') as f:
                    json.dump(orders, f, ensure_ascii=False, indent=4)
                logger.info(f"Заказ сохранён: {drink_name}, цена: {price}, дата: {selected_date}, телефон: {phone}")
                await update.message.reply_text(f"Ваш выбор ({drink_name}) записан! Цена: {price} рублей.")
            except Exception as e:
                logger.error(f"Ошибка записи в файл: {e}")
                await update.message.reply_text(f"Ошибка записи в файл: {e}")
   
    except Exception as e:
        logger.error(f"Ошибка при обработке комплексного обеда: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def handle_salad(update: Update, context: ContextTypes.DEFAULT_TYPE, salad_name: str):
    user_data = load_user_data()
    try:
            phone = context.user_data.get("phone_number")
            user = next((u for u in user_data["users"] if u["phone"] == phone), None)
            if phone is None:
                await update.message.reply_text("Ваш номер телефона не зарегистрирован, перезапустите бота!")
                return
            selected_date = context.user_data.get("selected_date")
            if selected_date is None:
                await update.message.reply_text("Выберите дату, прежде чем заказывать обед.")
                return
            selected_day_name = context.user_data.get("selected_day_name")
            address = user['address']
            if address is None:
                await update.message.reply_text("Вы не выбрали адрес, перезапустите бота!")
                return
            try:
                menu_data = pd.read_csv(MENU)
                salad_price = dict(zip(menu_data['Блюдо'], menu_data['Цена']))

                price = salad_price.get(salad_name)
                if price is None:
                    await update.message.reply_text(f"Цена для {salad_name} не найдена в меню.")
                    return
                try:
                    with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                        try:
                            orders = json.load(f)

                            if not isinstance(orders, list):
                                orders = []
                        except json.JSONDecodeError:

                            orders = []
                except FileNotFoundError:

                    orders = []

                new_order = {
                "Номер телефона": phone,
                "Дата": selected_date,
                "День недели": selected_day_name,
                "Обед": salad_name,
                "Цена": int(price),
                "Статус оплаты": "Не оплачено",
                "Адрес доставки": address,
                "Имя заказчика": user["name"],
                }
                try:
                    with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                        try:
                            orders = json.load(f)
                            if not isinstance(orders, list):
                                orders = []
                        except json.JSONDecodeError:
                            orders = []
                except FileNotFoundError:
                    orders = []

                orders.append(new_order)

                with open(ORDERS_JSON, 'w', encoding='utf-8') as f:
                    json.dump(orders, f, ensure_ascii=False, indent=4)
                logger.info(f"Заказ сохранён: {salad_name}, цена: {price}, дата: {selected_date}, телефон: {phone}")
                await update.message.reply_text(f"Ваш выбор ({salad_name}) записан! Цена: {price} рублей.")
            except Exception as e:
                logger.error(f"Ошибка записи в файл: {e}")
                await update.message.reply_text(f"Ошибка записи в файл: {e}")

    except Exception as e:
        logger.error(f"Ошибка при обработке комплексного обеда: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")

async def handle_complex_lunch(update: Update, context: ContextTypes.DEFAULT_TYPE, lunch_name: str):
    user_data = load_user_data()
    try:
        phone = context.user_data.get("phone_number")
        user = next((u for u in user_data["users"] if u["phone"] == phone), None)
        if phone is None:
            await update.message.reply_text("Ваш номер телефона не зарегистрирован, перезапустите бота!")
            return

        selected_date = context.user_data.get("selected_date")
        if selected_date is None:
            await update.message.reply_text("Выберите дату, прежде чем заказывать обед.")
            return
        selected_day_name = context.user_data.get("selected_day_name")
        address = user['address']
        if address is None:
            await update.message.reply_text("Вы не выбрали адрес, перезапустите бота!")
            return
        try:
            menu_data = pd.read_csv(MENU)
            lunch_prices = dict(zip(menu_data['Название'], menu_data['Цена']))

            price = lunch_prices.get(lunch_name)
            if price is None:
                await update.message.reply_text(f"Цена для {lunch_name} не найдена в меню.")
                return
            try:
                with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                    try:
                        orders = json.load(f)

                        if not isinstance(orders, list):
                            orders = []
                    except json.JSONDecodeError:

                        orders = []
            except FileNotFoundError:

                orders = []


            new_order = {
            "Номер телефона": phone,
            "Дата": selected_date,
            "День недели": selected_day_name,
            "Обед": lunch_name,
            "Цена": int(price),
            "Статус оплаты": "не оплачено",
            "Адрес доставки": address,
            "Имя заказчика": user["name"]
            }
            try:
                with open(ORDERS_JSON, 'r', encoding='utf-8') as f:
                    try:
                        orders = json.load(f)
                        if not isinstance(orders, list):
                            orders = []
                    except json.JSONDecodeError:
                        orders = []
            except FileNotFoundError:
                orders = []

            orders.append(new_order)

            with open(ORDERS_JSON, 'w', encoding='utf-8') as f:
                json.dump(orders, f, ensure_ascii=False, indent=4)
            logger.info(f"Заказ сохранён: {lunch_name}, цена: {price}, дата: {selected_date}, телефон: {phone}")
            await update.message.reply_text(f"Ваш выбор ({lunch_name}) записан! Цена: {price} рублей.")
        except Exception as e:
            logger.error(f"Ошибка записи в файл: {e}")
            await update.message.reply_text(f"Ошибка записи в файл: {e}")
    
    except Exception as e:
        logger.error(f"Ошибка при обработке комплексного обеда: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")


async def handle_payment_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    selected_option = update.message.text
    phone = context.user_data.get("phone_number")

    if phone is None:
        await update.message.reply_text("Ваш номер телефона не зарегистрирован. Перезапустите бота!")
        return

    if selected_option == "Оплатить картой💳":
        await pay(update, context)
        return

    elif selected_option == "Оплатить наличными":
        success, order_id = await move_orders_to_excel(phone, "Наличными")
        if success:
            await update.message.reply_text("Оплата наличными подтверждена. Ваш заказ перенесён в историю.")
            await show_main_menu(update, context)
        else:
            await update.message.reply_text("Ошибка при переносе заказа в историю.")

async def pay(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Создаёт платёж с суммой из корзины."""
    try:
        total_price = context.user_data.get("total_price", 0)
        if total_price == 0:
            await update.message.reply_text("Ваша корзина пуста, оплатить нечего.")
            return

        payment = Payment.create({
            "amount": {"value": f"{total_price}.00", "currency": "RUB"},
            "confirmation": {"type": "redirect", "return_url": "https://t.me/DirTasteBot"},
            "capture": True,
            "description": f"Оплата заказа на сумму {total_price} рублей"
        })

        context.user_data['payment_id'] = payment.id

        await update.message.reply_text(
            f'Платёж создан! Перейдите по ссылке({payment.confirmation.confirmation_url}) для оплаты.',
            parse_mode='Markdown'
        )

        asyncio.create_task(check_payment_status(update, context, payment.id))

    except Exception as e:
        logger.error(f'Ошибка при создании платежа: {str(e)}')
        await update.message.reply_text(f'Ошибка при создании платежа: {str(e)}')

async def check_payment_status(update: Update, context: ContextTypes.DEFAULT_TYPE, payment_id: str) -> None:
    try:
        start_time = datetime.now()
        while True:
            await asyncio.sleep(10)
            try:
                payment = Payment.find_one(payment_id)
                status = payment.status

                if status == 'succeeded':
                    phone = context.user_data.get("phone_number")
                    if phone:
                        success, order_id = await move_orders_to_excel(phone, "Картой")
                        if success:
                            await update.message.reply_text("Оплата прошла успешно! Ваш заказ перенесён в историю.")
                            await show_main_menu(update, context)
                            return
                        else:
                            await update.message.reply_text("Ошибка при переносе заказа в историю.")
                            return
                    return

                if status == 'pending':
                    # Проверяем, прошло ли 10 минут
                    if (datetime.now() - start_time).total_seconds() >= 600:
                        await cancel_payment(update, context, payment_id)
                        await update.message.reply_text("Время ожидания оплаты истекло. Платеж отменен.")
                        return

                if status == 'canceled':
                    await update.message.reply_text(f'Платеж {payment_id} отменен.')
                    return

                context.user_data['payment.status'] = status

            except Exception as e:
                logger.error(f'Ошибка при проверке статуса платежа: {str(e)}')
                await update.message.reply_text(f'Ошибка при проверке статуса платежа: {str(e)}')
                return

    except Exception as e:
        logger.error(f'Критическая ошибка в check_payment_status: {str(e)}')
        await update.message.reply_text('Произошла критическая ошибка. Пожалуйста, попробуйте позже.')

async def cancel_payment(update: Update, context: ContextTypes.DEFAULT_TYPE, payment_id: str) -> None:
    try:
        payment = Payment.find_one(payment_id)
        current_status = payment.status

        if current_status != 'succeeded':
            try:
                await clear_cart(update, context)
            except Exception as e:
                logger.error(f"Ошибка при очистке корзины: {e}")
                await update.message.reply_text("Ошибка при очистке корзины. Пожалуйста, попробуйте позже.")

    except Exception as e:
        logger.error(f"Ошибка при отмене платежа: {e}")

async def show_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):

    phone = context.user_data.get("phone_number")

    if phone is None:
        await update.message.reply_text("Ваш номер телефона не зарегистрирован. Перезапустите бота!")
        return ConversationHandler.END

    try:
        with open(ORDERS_JSON, "r", encoding="utf-8") as f:
            orders = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        await update.message.reply_text("Ошибка при загрузке заказов.")
        return ConversationHandler.END

    user_orders = [order for order in orders if order.get("Номер телефона") == phone]

    if not user_orders:
        await update.message.reply_text("Ваша корзина пуста.")
        return ConversationHandler.END

    from collections import defaultdict
    grouped_orders = defaultdict(lambda: {"Блюда": [], "Цена": 0, "День недели": ""})

    for order in user_orders:
        date = order["Дата"]
        grouped_orders[date]["Блюда"].append(order["Обед"])
        grouped_orders[date]["Цена"] += order["Цена"]
        grouped_orders[date]["День недели"] = order["День недели"]

    total_price = sum(details["Цена"] for details in grouped_orders.values())
    context.user_data["total_price"] = total_price  # Сохраняем сумму

    cart_message = "🛒 *Ваша корзина:*\n\n"
    for date, details in grouped_orders.items():
        cart_message += (
            f"📅 *Дата*: {date} ({details['День недели']})\n"
            f"🍽 *Состав заказа*: {', '.join(details['Блюда'])}\n"
            f"💰 *Цена*: {details['Цена']} рублей\n\n"
        )
    cart_message += f"💵 *Общая сумма*: {total_price} рублей"

    await update.message.reply_text(cart_message, parse_mode="Markdown")
    context.user_data["awaiting_comment"] = True
    keyboard = [["Пропустить комментарий"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    await update.message.reply_text("📝 Оставьте комментарий к заказу или нажмите 'Пропустить комментарий'.", reply_markup=reply_markup)

    return ENTER_COMMENT

async def handle_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        text = update.message.text.strip()
        if not context.user_data.get("awaiting_comment"):
            logger.warning("Comment handling called without awaiting_comment flag")
            return ConversationHandler.END

        if text == "Пропустить комментарий":
            comment = "Без комментария"
            await update.message.reply_text("Комментарий не добавлен. Переходим к выбору способа оплаты.")
        else:
            if len(text) > 500:  # Ограничение на длину комментария
                await update.message.reply_text("Комментарий слишком длинный. Максимальная длина - 500 символов.")
                return ENTER_COMMENT
            comment = text
            await update.message.reply_text(f"✅ Комментарий сохранён: {comment}. Теперь выберите способ оплаты.")

        context.user_data["comment"] = comment
        context.user_data["awaiting_comment"] = False

        phone = context.user_data.get("phone_number")
        if phone:
            try:
                with open(ORDERS_JSON, "r", encoding="utf-8") as f:
                    orders = json.load(f)
            except (FileNotFoundError, json.JSONDecodeError) as e:
                logger.error(f"Error loading orders: {e}")
                orders = []

            for order in orders:
                if order.get("Номер телефона") == phone:
                    order["Комментарий"] = comment

            try:
                with open(ORDERS_JSON, "w", encoding="utf-8") as f:
                    json.dump(orders, f, ensure_ascii=False, indent=4)
            except Exception as e:
                logger.error(f"Error saving orders with comment: {e}")

        return await show_payment_options(update, context)
    except Exception as e:
        logger.error(f"Error in handle_comment: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")
        return ConversationHandler.END

async def show_payment_options(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not context.user_data.get("total_price"):
            logger.warning("Payment options shown without total price")
            await update.message.reply_text("Ошибка: сумма заказа не найдена.")
            return ConversationHandler.END

        keyboard = [["Оплатить картой💳"], ["Оплатить наличными"], ["Назад 🔙"], ["Очистить корзину❌"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

        await update.message.reply_text("Выберите способ оплаты:", reply_markup=reply_markup)
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"Error in show_payment_options: {e}")
        await update.message.reply_text("Произошла ошибка. Пожалуйста, попробуйте снова.")
        return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Отмена ввода комментария.")
    return ConversationHandler.END

async def show_all_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("role") != "Администратор":
        await update.message.reply_text("У вас нет прав для использования этой функции.")
        return

    try:
        orders_df = pd.read_excel(ORDERS)
    except FileNotFoundError:
        await update.message.reply_text("Файл с заказами не найден.")
        return

    if orders_df.empty:
        await update.message.reply_text("Заказов пока нет.")
        return

    today = datetime.today().date()
    todaystr = today.strftime("%d.%m.%Y")
    today_orders = orders_df[orders_df['Дата'] == todaystr]
    if today_orders.empty:
        await update.message.reply_text("Заказов на сегодня нет.")
        return
    dish_count = {}
    dish_count_end = {}
    for index, row in today_orders.iterrows():
        address = row['Адрес доставки']
        dish = row['Обед']
        if address not in dish_count:
            dish_count[address] = {}
        if dish not in dish_count[address]:
            dish_count[address][dish] = 1
        else:
            dish_count[address][dish] += 1
        if dish not in dish_count_end:
            dish_count_end[dish] = 1
        else:
            dish_count_end[dish] += 1
    orders_text = "Список заказов на сегодня:\n\n"
    for address, dishes in dish_count.items():
        orders_text += f"Адрес доставки: {address}\n"
        for dish, count in dishes.items():
            orders_text += f"  - {dish}: {count}\n"
        orders_text += "\n"

    await update.message.reply_text(orders_text)

    orders_text = "Итого:\n"
    for dish, count in dish_count_end.items():
        orders_text += f"  - {dish}: {count}\n"

    await update.message.reply_text(orders_text)

def main():
    try:
        # Configure application with proper timeouts and update parameters
        application = (
            Application.builder()
            .token(TOKEN)
            .connect_timeout(30)
            .read_timeout(30)
            .write_timeout(30)
            .pool_timeout(30)
            .build()
        )

        registration_handler = ConversationHandler(
            entry_points=[MessageHandler(filters.CONTACT, start)],
            states={
                CHOOSE_ADDRESS: [CallbackQueryHandler(choose_address)],
                ENTER_NAME: [MessageHandler(filters.TEXT, enter_name)],
            },
            fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        )

        broadcast_handler = ConversationHandler(
            entry_points=[MessageHandler(filters.Regex("^Сообщить всем$"), broadcast_start)],
            states={
                BROADCAST_MESSAGE: [MessageHandler(filters.TEXT, broadcast_message)],
            },
            fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        )

        address_handler = ConversationHandler(
            entry_points=[MessageHandler(filters.Regex("^Добавить адрес доставки$"), add_address_start)],
            states={
                ADD_ADDRESS: [MessageHandler(filters.TEXT, add_address)],
            },
            fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        )
        comment_handler = ConversationHandler(
            entry_points=[MessageHandler(filters.Regex("^Корзина 🗑$"), show_cart)],
            states={
                ENTER_COMMENT: [MessageHandler(filters.TEXT, handle_comment)]
            },
            fallbacks=[CommandHandler("cancel", cancel)],
        )

        # Add handlers
        application.add_handler(CommandHandler("start", under_start))
        application.add_handler(registration_handler)
        application.add_handler(broadcast_handler)
        application.add_handler(address_handler)
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_buttons))
        application.add_handler(CallbackQueryHandler(handle_menu_and_lunch))
        application.add_handler(CallbackQueryHandler(button_callback))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_buttons))
        application.add_handler(MessageHandler(filters.Regex("^Корзина 🗑$"), show_cart))
        application.add_handler(comment_handler)

        # Start the bot with proper polling configuration
        application.run_polling(
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True
        )
    except Exception as e:
        logger.error(f"Ошибка в main: {e}")
        raise

if __name__ == "__main__":
    main()

